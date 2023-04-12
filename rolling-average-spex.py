import requests
import datetime
from datetime import date, timedelta
from bs4 import BeautifulSoup
import pandas as pd
import lxml
import openpyxl
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
import matplotlib.pyplot as plt

def remove_duplicate_rows(matching_sheets):
    """
    Function to remove duplicate rows from the 'matching_sheets' dataframe based on the 'Sheet' column.

    Parameters:
    matching_sheets (pandas.DataFrame): The dataframe containing the matching sheets and their speX values.

    Returns:
    pandas.DataFrame: The updated dataframe with duplicate rows removed.
    """
    # remove duplicate rows based on the 'Sheet' column
    matching_sheets = matching_sheets.drop_duplicates(subset=['Sheet'])

    # reset the index of the dataframe
    matching_sheets = matching_sheets.reset_index(drop=True)

    return matching_sheets


def parse_array_from_fangraphs_html(start_date,end_date):
    """
    Take a HTML stats page from fangraphs and parse it out to a dataframe.
    """
    # parse input
    PITCHERS_URL = "https://www.fangraphs.com/leaders.aspx?pos=all&stats=pit&lg=all&qual=0&type=c%2C13%2C7%2C8%2C120%2C121%2C331%2C105%2C111%2C24%2C19%2C14%2C329%2C324%2C45%2C122%2C6%2C42%2C43%2C328%2C330%2C322%2C323%2C326%2C332&season=2021&month=1000&season1=2015&ind=0&team=&rost=&age=&filter=&players=&startdate={}&enddate={}&page=1_2000".format(start_date, end_date)
    PITCHERS_URL = "https://www.fangraphs.com/leaders.aspx?pos=all&stats=pit&lg=all&qual=0&type=c,13,7,8,120,121,331,105,111,24,19,14,329,324,45,122,6,42,43,328,330,322,323,326,332,31,30,29&season=2021&month=1000&season1=2015&ind=0&team=&rost=&age=&filter=&players=&startdate={}&enddate={}&page=1_2000".format(start_date, end_date)

    # request the data
    pitchers_html = requests.get(PITCHERS_URL).text
    soup = BeautifulSoup(pitchers_html, "lxml")
    table = soup.find("table", {"class": "rgMasterTable"})
    
    # get headers
    headers_html = table.find("thead").find_all("th")
    headers = []
    for header in headers_html:
        headers.append(header.text)

    # get rows
    rows = []
    rows_html = table.find("tbody").find_all("tr")
    for row in rows_html:
        row_data = []
        for cell in row.find_all("td"):
            row_data.append(cell.text)
        rows.append(row_data)
    
    return pd.DataFrame(rows, columns = headers)

def calc_speX(df, IP_limit):
    work_speX = df[['Name','Team', 'IP', 'G', 'GS','K%','BB%','SO','BB','TBF',
                    'Events','Barrels','CSW%','O-Swing%','Zone%',
                    'Pitches', 'Balls', 'Strikes']].copy()
    
    allowed_chars = set('0123456789.%')
    work_speX = work_speX[work_speX['Zone%'].apply(lambda x: set(x).issubset(allowed_chars))]
    
   
    work_speX.insert(loc=5, column='K-BB%', value=['' for i in range(work_speX.shape[0])])
    work_speX.insert(loc=6, column='pCRA', value=['' for i in range(work_speX.shape[0])])
    work_speX.insert(loc=17, column='O-Sw%+Z%', value=['' for i in range(work_speX.shape[0])])
    work_speX.insert(loc=18, column='speX', value=['' for i in range(work_speX.shape[0])])

    work_speX['K%'] = work_speX['K%'].map(lambda x: x.rstrip('%'))
    work_speX['BB%'] = work_speX['BB%'].map(lambda x: x.rstrip('%'))
    work_speX['CSW%'] = work_speX['CSW%'].map(lambda x: x.rstrip('%'))
    work_speX['O-Swing%'] = work_speX['O-Swing%'].map(lambda x: x.rstrip('%'))
    work_speX['Zone%'] = work_speX['Zone%'].map(lambda x: x.rstrip('%'))
    work_speX = work_speX.replace(r'^\s*$', 0, regex=True)    

    
    work_speX['K%'] = work_speX['K%'].astype(float)
    work_speX['BB%'] = work_speX['BB%'].astype(float)
    work_speX['CSW%'] = work_speX['CSW%'].astype(float)
    work_speX['O-Swing%'] = work_speX['O-Swing%'].astype(float)
    work_speX['SO'] = work_speX['SO'].astype(float)
    work_speX['TBF'] = work_speX['TBF'].astype(float)
    work_speX['BB'] = work_speX['BB'].astype(float)
    work_speX['Barrels'] = work_speX['Barrels'].astype(float)
    work_speX['Events'] = work_speX['Events'].astype(float)
    work_speX['IP'] = work_speX['IP'].astype(float)
    work_speX['Pitches'] = work_speX['Pitches'].astype(float)
    work_speX['Balls'] = work_speX['Balls'].astype(float)
    work_speX['Zone%'] = work_speX['Zone%'].astype(float)
    work_speX['Zone%'] = work_speX['Zone%'].fillna(0)


    

    work_speX['K-BB%'] = work_speX['K%'] - work_speX['BB%']
    work_speX['CSW-BB%'] = work_speX['CSW%'] - work_speX['BB%']
    work_speX['CSW-B%'] = work_speX['CSW%'] - (work_speX['Balls']/work_speX['Pitches'])*100
    work_speX['CSW-W%'] = work_speX['CSW%'] - (work_speX['BB']/work_speX['Pitches'])*100
    work_speX['O-Sw%+Z%'] = work_speX['O-Swing%'] + work_speX['Zone%']
    work_speX['pCRA'] = (-8.89 * ((work_speX['SO'] + 10.66) / (work_speX['TBF'] + 10.66 + 37.53)) + 8.03 * ((work_speX['BB'] + 9.58) /(work_speX['TBF'] + 9.58 + 116.2)) + 5.54 * ((work_speX['Barrels'] + 13.33) / (work_speX['Events'] + 13.33 + 191.5)) + 106 * ((work_speX['Barrels'] + 13.33) / (work_speX['Events'] + 13.33 + 191.5)) * ((work_speX['Barrels'] + 13.33) / (work_speX['Events'] + 13.33 + 191.5))) + 4.55
    work_speX['speX'] = (11 * work_speX['K-BB%'] + 165 * (10-work_speX['pCRA']) + 7 * work_speX['CSW%'] + work_speX['O-Sw%+Z%']) * 0.049

    speX = work_speX.drop(['K%', 'BB%', 'SO', 'BB', 'TBF', 'Events', 'Barrels', 'Strikes'], axis = 1)


    speX['K-BB%'] = speX['K-BB%'].round(decimals=2)
    speX['pCRA'] = speX['pCRA'].round(decimals=2)
    speX['CSW%'] = speX['CSW%'].round(decimals=2)
    speX['O-Swing%'] = speX['O-Swing%'].round(decimals=2)
    speX['Zone%'] = speX['Zone%'].round(decimals=2)
    speX['O-Sw%+Z%'] = speX['O-Sw%+Z%'].round(decimals=2)
    speX['speX'] = speX['speX'].round(decimals=2)


    speX = speX.sort_values(by=['speX'], ascending=False)

    speX = speX[speX['IP'] >= IP_limit]
    speX = speX.sort_values(by=['speX'], ascending=False)
    
    return speX


def date_range(start_date, end_date):
    for n in range(int((pd.to_datetime(end_date) - pd.to_datetime(start_date)).days)+1):
        yield pd.to_datetime(start_date) + pd.DateOffset(n)


# define start and end dates
sdate = '2023-03-30'
enddate = '2023-04-12'
IP = 0 
speX_file = 'speX-rolling.xlsx'
# convert start and end dates to datetime objects
start_date = pd.to_datetime(sdate)
end_date = pd.to_datetime(enddate)

writer = pd.ExcelWriter(speX_file, engine='openpyxl') 

# loop through date range
for single_date in date_range(sdate, enddate):
    # calculate dataframe for each date
    #df = calc_speX(df, IP_limit)
    temp_date = single_date.strftime('%Y-%m-%d')
    print(f"Results for {temp_date}:")
    #print(df)
    speX = parse_array_from_fangraphs_html(sdate, temp_date)
    result = calc_speX(speX, IP)
    result.to_excel(writer, sheet_name=temp_date)

speX = parse_array_from_fangraphs_html(sdate, enddate)
speX_done = calc_speX(speX, IP)

speX_done.to_excel(writer, sheet_name='full')

writer.save()
writer.close()

# Open the Excel file
workbook = openpyxl.load_workbook(speX_file)

# Get the worksheet you want to move
worksheet = workbook['full']

# Get the current sheet index of the worksheet
sheet_index = workbook.index(worksheet)

# Move the worksheet to the beginning
workbook.move_sheet(worksheet, offset=-sheet_index)

# Iterate through each worksheet in the workbook
for worksheet in workbook.worksheets:
# Auto-fit all columns in the worksheet
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


# Save the changes to the Excel file
workbook.save(speX_file)
workbook.close()

wb = load_workbook(filename=speX_file)

# get a list of all the sheet names
sheet_names = wb.sheetnames

# create a dataframe to store matching sheets and their speX values
matching_sheets = pd.DataFrame(columns=['Sheet', 'speX'])

# Get the full sheet
full_sheet = wb['full']

# Get the headers
headers = [cell.value for cell in full_sheet[1]]

# Create a dictionary of header indices
header_dict = {header: index for index, header in enumerate(headers)}

# Get the data as a list of lists
data = [[cell.value for cell in row] for row in full_sheet.iter_rows(min_row=2)]

# Convert the data to a pandas dataframe
df = pd.DataFrame(data, columns=headers)

# Ask the user for a player name
player_name = input("Enter a player name: ")

# Search for a match in the "Name" column
name_column = df['Name']
match_mask = name_column.str.contains(player_name, case=False)
if match_mask.any():
    print("Match found:")
    matching_row = df[match_mask].iloc[0]
    print("IP:", matching_row['IP'])
    print("G:", matching_row['G'])
    print("GS:", matching_row['GS'])
    print("K-BB%:", matching_row['K-BB%'])
    print("pCRA:", matching_row['pCRA'])
    print("CSW%:", matching_row['CSW%'])
    print("O-Swing%:", matching_row['O-Swing%'])
    print("Zone%:", matching_row['Zone%'])
    print("O-Sw%+Z%:", matching_row['O-Sw%+Z%'])
    print("speX:", matching_row['speX'])

    # iterate through each sheet and search for the player name
    for sheet_name in sheet_names:
        # load sheet into temporary dataframe
        df = pd.read_excel(speX_file, sheet_name)
        
        # check if 'Name' column exists in the temporary dataframe
        if 'Name' in df.columns:
            # search for the player name in the 'Name' column
            name_column = df['Name']
            for i, row_value in name_column.items():
                if isinstance(row_value, str) and fuzz.ratio(player_name.lower(), row_value.lower()) > 75:
                    # add matching sheet and speX value to the dataframe
                    matching_sheets = matching_sheets.append({'Sheet': sheet_name, 'speX': df.iloc[i]['speX']}, ignore_index=True)

    # print the matching sheets and their speX values
    print(matching_sheets)


else:
    print("No match found.")

wb.close()

# Drop duplicates, keeping the first occurrence of each
matching_sheets.drop_duplicates(subset=['Sheet'], keep='first', inplace=True)

# Drop rows where 'Sheet' == 'full'
matching_sheets = matching_sheets[matching_sheets['Sheet'] != 'full']

# Drop duplicates for the column 'speX', keeping the first occurrence of each
matching_sheets.drop_duplicates(subset=['speX'], keep='first', inplace=True)


# Reset the index
matching_sheets.reset_index(drop=True, inplace=True)

# Filter dataframe by date range
#start_date = datetime.datetime.strptime(sdate, '%Y-%m-%d')
#end_date = datetime.datetime.strptime(enddate, '%Y-%m-%d')
#matching_sheets = matching_sheets[(matching_sheets['Sheet'] >= start_date) & (matching_sheets['Sheet'] <= end_date)]

# Create graph
#fig, ax = plt.subplots(figsize=(10, 6))
#ax.plot(matching_sheets['Sheet'], matching_sheets['speX'])
#ax.set_ylim([40, 100])
#ax.set_xlabel('Game Dates')
#ax.set_ylabel('Accumulated speX')
#ax.set_title(player_name + '\n' + 'speX through games from ' + sdate + ' to ' + enddate)
#plt.show()

# Create graph
fig, ax = plt.subplots(figsize=(10, 6))
ax.plot(range(1, len(matching_sheets) + 1), matching_sheets['speX'])
ax.set_xlim([1, len(matching_sheets)])
ax.set_ylim([40, 100])
ax.set_xticks(range(1, len(matching_sheets) + 1))
ax.set_xticklabels([i for i in range(1, len(matching_sheets) + 1)])

ax.set_xlabel('Games')
ax.set_ylabel('Accumulated speX')
ax.set_title(player_name + '\n' + 'speX through games from ' + sdate + ' to ' + enddate)
plt.show()
